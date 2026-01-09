import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox

# -----------------------------
# Function to generate report
# -----------------------------
def generate_report(input_file, output_file):
    try:
        # -----------------------------
        # Read input file
        # -----------------------------
        file_ext = input_file.split(".")[-1].lower()

        if file_ext in ["xlsx", "xls"]:
            df = pd.read_excel(input_file)
        elif file_ext == "csv":
            df = pd.read_csv(input_file)
        else:
            raise ValueError("Unsupported file format. Use CSV or Excel.")

        # -----------------------------
        # Select columns and rename
        # -----------------------------
        column_map = {
            "Work date": "work date",
            "Shift": "shift",
            "Machine": "machine",
            "Q`ty": "worked Q'ty",
            "WPCS Qty": "WPC qty"
        }

        df_selected = df[list(column_map.keys())].rename(columns=column_map)

        # -----------------------------
        # Convert work date to datetime safely
        # -----------------------------
        df_selected['work date'] = pd.to_datetime(df_selected['work date'], errors='coerce')
        df_selected = df_selected.dropna(subset=['work date'])

        # -----------------------------
        # Filter machines A01 to A38
        # -----------------------------
        allowed_machines = [f"A{i:02d}" for i in range(1, 39)]
        df_filtered = df_selected[df_selected["machine"].isin(allowed_machines)]

        # -----------------------------
        # Find last date per machine
        # -----------------------------
        last_dates = df_filtered.groupby("machine")["work date"].max().reset_index()
        last_dates.rename(columns={"work date": "last_work_date"}, inplace=True)

        # -----------------------------
        # Sum all qty per machine
        # -----------------------------
        sum_qty = df_filtered.groupby("machine")[["worked Q'ty", "WPC qty"]].sum().reset_index()

        # -----------------------------
        # Combine sum with last date
        # -----------------------------
        summary_df = pd.merge(last_dates, sum_qty, on="machine")

        # -----------------------------
        # Add WPCS % column (rounded to 2 decimals)
        # -----------------------------
        summary_df['WPCS %'] = summary_df.apply(
            lambda row: round((row['WPC qty'] / row["worked Q'ty"]) * 100, 2) if row["worked Q'ty"] != 0 else 0, axis=1
        )

        # -----------------------------
        # Remove rows where WPCS % is below 20%
        # -----------------------------
        summary_df = summary_df[summary_df['WPCS %'] >= 20]

        # -----------------------------
        # Create total summary table
        # -----------------------------
        total_worked_qty = summary_df["worked Q'ty"].sum()
        total_wpc_qty = summary_df["WPC qty"].sum()
        total_wpcs_percent = round((total_wpc_qty / total_worked_qty) * 100, 2) if total_worked_qty != 0 else 0

        summary_total = pd.DataFrame([{
            "Total worked Q'ty": total_worked_qty,
            "Total WPC qty": total_wpc_qty,
            "Total WPCS %": total_wpcs_percent
        }])

        # -----------------------------
        # Write both tables to a single sheet
        # -----------------------------
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Report', index=False, startrow=0)
            start_row_total = len(summary_df) + 2
            summary_total.to_excel(writer, sheet_name='Report', index=False, startrow=start_row_total)

        # -----------------------------
        # Styling with openpyxl
        # -----------------------------
        wb = load_workbook(output_file)
        ws = wb['Report']

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        center_align = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    value_length = len(str(cell.value))
                    if value_length > max_length:
                        max_length = value_length
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Style headers of machine-level table
        for col in range(1, len(summary_df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Style data of machine-level table
        for row in ws.iter_rows(min_row=2, max_row=1 + len(summary_df), min_col=1, max_col=len(summary_df.columns)):
            for cell in row:
                cell.border = border
                cell.alignment = center_align

        # Style headers of total summary table
        start_total = len(summary_df) + 3
        for col in range(1, len(summary_total.columns) + 1):
            cell = ws.cell(row=start_total, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Style data of total summary table
        for row in ws.iter_rows(min_row=start_total + 1, max_row=start_total + len(summary_total), min_col=1, max_col=len(summary_total.columns)):
            for cell in row:
                cell.border = border
                cell.alignment = center_align

        wb.save(output_file)

        messagebox.showinfo("Success", f"✅ Excel report created successfully!\nSaved to: {output_file}")

    except Exception as e:
        messagebox.showerror("Error", f"❌ An error occurred:\n{e}")

# -----------------------------
# GUI Setup
# -----------------------------
def select_input_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv")])
    if file_path:
        input_entry.delete(0, tk.END)
        input_entry.insert(0, file_path)

def select_output_file():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        output_entry.delete(0, tk.END)
        output_entry.insert(0, file_path)

def run_report():
    input_file_path = input_entry.get()
    output_file_path = output_entry.get()
    if not input_file_path or not output_file_path:
        messagebox.showwarning("Warning", "Please select both input and output files.")
        return
    generate_report(input_file_path, output_file_path)

# GUI window
root = tk.Tk()
root.title("Komax Report Generator")

tk.Label(root, text="Input File:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
input_entry = tk.Entry(root, width=50)
input_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=select_input_file).grid(row=0, column=2, padx=10, pady=10)

tk.Label(root, text="Output File:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
output_entry = tk.Entry(root, width=50)
output_entry.grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=select_output_file).grid(row=1, column=2, padx=10, pady=10)

tk.Button(root, text="Generate Report", command=run_report, bg="#4F81BD", fg="white", font=("Arial", 12, "bold")).grid(row=2, column=0, columnspan=3, pady=20)

root.mainloop()
